"""BOM item management service with supplier integration and export."""

from __future__ import annotations

import csv
import logging
from decimal import Decimal
from pathlib import Path
from typing import Optional
from uuid import UUID

from bom_manager.core.exceptions import (
    ExportError,
    ItemNotFoundError,
    SupplierLookupError,
    VersionNotFoundError,
)
from bom_manager.core.models import BOMItem, BOMSummary, PriceBreak
from bom_manager.storage.base import StorageProtocol
from bom_manager.suppliers.base import PartDetail, PartResult, SupplierError, SupplierProtocol

log = logging.getLogger(__name__)

_EXPORT_DIR = Path("exports")

# CSV column headers, in output order
_CSV_HEADERS = [
    "Reference",
    "Part Name",
    "MPN",
    "Supplier",
    "Supplier PN",
    "Quantity",
    "Unit Price",
    "Total Price",
    "URL",
]


class BOMService:
    """
    Manages BOM items: searching suppliers, adding/removing parts,
    updating quantities, computing costs, and exporting.

    Parameters
    ----------
    storage:
        Persistence backend satisfying ``StorageProtocol``.
    supplier:
        Supplier adapter satisfying ``SupplierProtocol`` (e.g. ``LCSCSupplier``).
    export_dir:
        Directory where exported files are written.  Created on demand.
        Defaults to ``exports/`` relative to the working directory.
    """

    def __init__(
        self,
        storage: StorageProtocol,
        supplier: SupplierProtocol,
        export_dir: Path = _EXPORT_DIR,
    ) -> None:
        self._storage = storage
        self._supplier = supplier
        self._export_dir = export_dir

    # ------------------------------------------------------------------
    # Supplier search (separate from add so callers can present choices)
    # ------------------------------------------------------------------

    def search_parts(self, query: str) -> list[PartResult]:
        """
        Search the supplier for *query* and return ranked results.

        Call this first to let the user pick a result, then pass the chosen
        ``supplier_pn`` to :meth:`add_part`.

        Returns an empty list when nothing matches.
        May raise ``SupplierError`` on network failures.
        """
        return self._supplier.search(query)

    # ------------------------------------------------------------------
    # Add part
    # ------------------------------------------------------------------

    def add_part(
        self,
        version_id: UUID,
        user_part_name: str,
        quantity: int,
        reference_designator: str,
        *,
        supplier_pn: Optional[str] = None,
    ) -> BOMItem:
        """
        Add a part to a BOM version.

        Workflow
        --------
        1. If *supplier_pn* is ``None``: search the supplier using
           *user_part_name* as the query and auto-select the top result.
           To let the user pick, call :meth:`search_parts` first, present
           the list, then pass the chosen ``supplier_pn`` here.
        2. Fetch full part detail (MPN, stock, price breaks) from the supplier.
        3. Resolve the correct price break tier for *quantity*.
        4. Persist the ``BOMItem`` and return it.

        Parameters
        ----------
        version_id:
            UUID of the target BOM version.
        user_part_name:
            Human-readable label the user typed (e.g. ``"100nF 0402 cap"``).
        quantity:
            Number of units required.
        reference_designator:
            PCB reference designator(s), e.g. ``"C1"`` or ``"C3,C4"``.
        supplier_pn:
            Specific supplier catalogue number.  When omitted the top search
            result is used automatically.

        Raises
        ------
        VersionNotFoundError
            If the version does not exist.
        SupplierLookupError
            If the search returns no results (only when *supplier_pn* is None).
        SupplierError
            On supplier network or parse failures.
        """
        # Guard: version must exist
        if self._storage.get_version(version_id) is None:
            raise VersionNotFoundError(f"Version {version_id} not found")

        # Resolve supplier PN via search when not explicitly provided
        if supplier_pn is None:
            results = self._supplier.search(user_part_name)
            if not results:
                raise SupplierLookupError(
                    f"No supplier results found for {user_part_name!r}"
                )
            supplier_pn = results[0].supplier_pn
            log.debug(
                "add_part: auto-selected %r from search results for %r",
                supplier_pn, user_part_name,
            )

        # Fetch full detail (price breaks, stock, MPN, …)
        try:
            detail: PartDetail = self._supplier.get_part(supplier_pn)
        except SupplierError:
            raise
        except Exception as exc:
            raise SupplierLookupError(
                f"Unexpected error fetching {supplier_pn!r}: {exc}"
            ) from exc

        item = _build_item(
            version_id=version_id,
            user_part_name=user_part_name,
            quantity=quantity,
            reference_designator=reference_designator,
            detail=detail,
            supplier_name=self._supplier.name,
        )
        saved = self._storage.add_item(item)
        log.info(
            "Added %r (%s / %s) × %d  @  %s USD  to version %s",
            user_part_name,
            saved.matched_mpn,
            saved.supplier_part_number,
            quantity,
            saved.unit_price,
            version_id,
        )
        return saved

    # ------------------------------------------------------------------
    # Remove part
    # ------------------------------------------------------------------

    def remove_part(self, version_id: UUID, item_id: UUID) -> None:
        """
        Remove a BOM item from a version.

        Raises
        ------
        ItemNotFoundError
            If no item with *item_id* exists in *version_id*.
        """
        item = self._get_item(version_id, item_id)
        self._storage.remove_item(item.id)
        log.info("Removed item %s (%r) from version %s", item_id, item.user_part_name, version_id)

    # ------------------------------------------------------------------
    # Update quantity
    # ------------------------------------------------------------------

    def update_quantity(
        self,
        version_id: UUID,
        item_id: UUID,
        new_qty: int,
    ) -> BOMItem:
        """
        Change the quantity of a BOM item and recalculate its unit price.

        The correct price break tier is chosen automatically based on the
        new quantity using the price breaks already stored on the item.

        Raises
        ------
        ItemNotFoundError
            If the item does not exist in the version.
        ValueError
            If *new_qty* is not a positive integer.
        """
        if new_qty < 1:
            raise ValueError(f"Quantity must be >= 1, got {new_qty}")

        item = self._get_item(version_id, item_id)

        # Build updated item with the new quantity; effective_unit_price()
        # uses stored price_breaks to pick the correct tier automatically.
        updated = item.model_copy(update={"quantity": new_qty})
        new_unit_price = updated.effective_unit_price()
        new_total = (
            new_unit_price * Decimal(new_qty)
            if new_unit_price is not None
            else None
        )
        updated = updated.model_copy(
            update={"unit_price": new_unit_price, "total_price": new_total}
        )

        saved = self._storage.update_item(updated)
        log.info(
            "Updated item %s quantity %d → %d  (unit price now %s USD)",
            item_id, item.quantity, new_qty, saved.unit_price,
        )
        return saved

    # ------------------------------------------------------------------
    # BOM summary
    # ------------------------------------------------------------------

    def get_bom(self, version_id: UUID) -> BOMSummary:
        """
        Return a full BOM summary for a version.

        Total cost is calculated from each item's price breaks and quantity.
        Items with no price data contribute $0 to the total.

        Raises
        ------
        VersionNotFoundError
            If the version does not exist.
        """
        if self._storage.get_version(version_id) is None:
            raise VersionNotFoundError(f"Version {version_id} not found")

        items = self._storage.list_items_by_version(version_id)
        return BOMSummary.from_items(version_id, items)

    # ------------------------------------------------------------------
    # Export
    # ------------------------------------------------------------------

    def export_bom(
        self,
        version_id: UUID,
        format: str = "csv",
        *,
        output_dir: Optional[Path] = None,
        filename: Optional[str] = None,
    ) -> Path:
        """
        Export the BOM for *version_id* to a file.

        Parameters
        ----------
        version_id:
            UUID of the version to export.
        format:
            Export format.  ``"csv"`` or ``"xlsx"``.
        output_dir:
            Directory to write the file into.  Defaults to ``exports/``.
        filename:
            Override the generated filename (without extension).

        Returns
        -------
        Path
            Absolute path to the written file.

        Raises
        ------
        ValueError
            If *format* is not supported.
        ExportError
            If the file cannot be written.
        """
        if format not in ("csv", "xlsx"):
            raise ValueError(
                f"Unsupported export format {format!r}. Supported: 'csv', 'xlsx'"
            )

        summary = self.get_bom(version_id)
        out_dir = output_dir or self._export_dir
        stem = filename or f"bom_{version_id}"
        path = (out_dir / f"{stem}.{format}").resolve()

        try:
            out_dir.mkdir(parents=True, exist_ok=True)
            if format == "csv":
                _write_csv(path, summary)
            else:
                _write_xlsx(path, summary)
        except OSError as exc:
            raise ExportError(f"Could not write export to {path}: {exc}") from exc

        log.info("Exported BOM for version %s → %s", version_id, path)
        return path

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _get_item(self, version_id: UUID, item_id: UUID) -> BOMItem:
        """Return the item or raise ItemNotFoundError."""
        items = self._storage.list_items_by_version(version_id)
        item = next((i for i in items if i.id == item_id), None)
        if item is None:
            raise ItemNotFoundError(
                f"Item {item_id} not found in version {version_id}"
            )
        return item


# ---------------------------------------------------------------------------
# Module-level helpers
# ---------------------------------------------------------------------------


def _best_unit_price_for_qty(
    detail: PartDetail, quantity: int
) -> Optional[Decimal]:
    """
    Pick the lowest unit price from *detail.price_breaks* that applies at
    *quantity*.  Falls back to the cheapest available break when quantity
    is below all thresholds (e.g. sampling one unit of a reel part).
    """
    if not detail.price_breaks:
        return None
    eligible = [pb for pb in detail.price_breaks if pb.min_quantity <= quantity]
    if eligible:
        return min(eligible, key=lambda pb: pb.unit_price).unit_price
    # quantity is below the minimum break — return the first (highest) price
    return min(detail.price_breaks, key=lambda pb: pb.min_quantity).unit_price


def _build_item(
    *,
    version_id: UUID,
    user_part_name: str,
    quantity: int,
    reference_designator: str,
    detail: PartDetail,
    supplier_name: str,
) -> BOMItem:
    """Construct a ``BOMItem`` from a ``PartDetail`` + context."""
    # Convert supplier PriceBreakInfo → domain PriceBreak
    price_breaks = [
        PriceBreak(
            min_quantity=pb.min_quantity,
            unit_price=pb.unit_price,
        )
        for pb in detail.price_breaks
    ]
    unit_price = _best_unit_price_for_qty(detail, quantity)
    total_price = unit_price * Decimal(quantity) if unit_price is not None else None

    return BOMItem(
        version_id=version_id,
        reference_designator=reference_designator,
        user_part_name=user_part_name,
        matched_mpn=detail.mpn or None,
        supplier=supplier_name,
        supplier_part_number=detail.supplier_pn or None,
        supplier_url=detail.url or None,
        quantity=quantity,
        unit_price=unit_price,
        price_breaks=price_breaks,
        total_price=total_price,
    )


def _item_row(item: BOMItem) -> list:
    """
    Return a flat list of cell values for one BOMItem.

    Column order matches _CSV_HEADERS:
    Reference | Part Name | MPN | Supplier | Supplier PN |
    Quantity  | Unit Price | Total Price | URL
    """
    unit_price = item.effective_unit_price()
    total = item.calculate_total()
    return [
        item.reference_designator,                          # 0 Reference
        item.user_part_name,                                # 1 Part Name
        item.matched_mpn or "",                             # 2 MPN
        item.supplier or "",                                # 3 Supplier
        item.supplier_part_number or "",                    # 4 Supplier PN
        item.quantity,                                      # 5 Quantity
        float(unit_price) if unit_price is not None else "",  # 6 Unit Price
        float(total) if total is not None else "",          # 7 Total Price
        item.supplier_url or "",                            # 8 URL
    ]


def _write_csv(path: Path, summary: BOMSummary) -> None:
    """Write *summary* to a CSV file at *path*."""
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(_CSV_HEADERS)
        for item in summary.items:
            row = _item_row(item)
            # Format numeric prices as strings for CSV readability
            row[6] = f"{row[6]:.4f}" if row[6] != "" else ""
            row[7] = f"{row[7]:.4f}" if row[7] != "" else ""
            writer.writerow(row)


def _write_xlsx(path: Path, summary: BOMSummary) -> None:
    """Write *summary* to an Excel workbook at *path*."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOM"

    # Header row
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    for col, header in enumerate(_CSV_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, item in enumerate(summary.items, 2):
        for col_idx, value in enumerate(_item_row(item), 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

    # Total row  (col 7 = Unit Price label, col 8 = Total Price value)
    total_row = len(summary.items) + 2
    ws.cell(row=total_row, column=7, value="TOTAL")
    ws.cell(row=total_row, column=7).font = Font(bold=True)
    ws.cell(row=total_row, column=8, value=float(summary.total_cost))
    ws.cell(row=total_row, column=8).font = Font(bold=True)

    wb.save(path)
